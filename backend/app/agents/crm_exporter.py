"""
CRMExporter — no es un agente LLM.
Toma los datos consolidados del pipeline y genera el archivo Excel final con openpyxl.
"""

import os
from datetime import datetime
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

# ── Paleta de colores ──────────────────────────────────────────────────────────
_HEADER_BG = "2E4057"   # azul oscuro SEÑAL
_HEADER_FG = "FFFFFF"
_HOT_BG = "D9EAD3"      # verde claro — caliente ≥ 8
_WARM_BG = "FFF2CC"     # amarillo claro — tibio 6–7.9


def _header_style() -> tuple[Font, PatternFill, Alignment]:
    font = Font(bold=True, color=_HEADER_FG)
    fill = PatternFill("solid", fgColor=_HEADER_BG)
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align


def _row_fill(score: float | None) -> PatternFill | None:
    if score is None:
        return None
    if score >= 8:
        return PatternFill("solid", fgColor=_HOT_BG)
    if score >= 6:
        return PatternFill("solid", fgColor=_WARM_BG)
    return None


def _write_headers(ws, headers: list[str], col_widths: list[int]) -> None:
    font, fill, align = _header_style()
    for col_idx, (h, w) in enumerate(zip(headers, col_widths), start=1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = font
        cell.fill = fill
        cell.alignment = align
        ws.column_dimensions[get_column_letter(col_idx)].width = w
    ws.row_dimensions[1].height = 30


def _set_cell(ws, row: int, col: int, value, wrap: bool = False, score: float | None = None) -> None:
    cell = ws.cell(row=row, column=col, value=value)
    cell.alignment = Alignment(vertical="top", wrap_text=wrap)
    if score is not None:
        fill = _row_fill(score)
        if fill:
            cell.fill = fill


# ── Pestaña 1: Empresas ────────────────────────────────────────────────────────
_EMP_HEADERS = [
    "ID", "Empresa", "NIT", "Sector", "Subsector", "Ciudad",
    "Tamaño estimado", "Web", "LinkedIn empresa",
    "Score (1-10)", "Clasificación", "Plan sugerido", "Fuente", "Contexto",
]
_EMP_WIDTHS = [8, 35, 18, 18, 18, 18, 20, 30, 30, 12, 14, 16, 20, 50]


def _build_empresas(ws, companies: list[dict], scores_map: dict) -> None:
    _write_headers(ws, _EMP_HEADERS, _EMP_WIDTHS)
    for row_i, emp in enumerate(companies, start=2):
        eid = emp.get("id_empresa", "")
        score_info = scores_map.get(eid, {})
        score = score_info.get("score")
        row = [
            eid,
            emp.get("nombre_empresa"),
            emp.get("nit"),
            emp.get("sector"),
            emp.get("subsector"),
            emp.get("ciudad"),
            emp.get("tamano_estimado"),
            emp.get("web"),
            emp.get("linkedin_empresa"),
            score,
            score_info.get("clasificacion"),
            score_info.get("plan_sugerido"),
            emp.get("fuente"),
            emp.get("contexto_adicional"),
        ]
        for col_i, val in enumerate(row, start=1):
            wrap = col_i == len(_EMP_HEADERS)  # Columna Contexto
            _set_cell(ws, row_i, col_i, val, wrap=wrap, score=score)
        if wrap:
            ws.row_dimensions[row_i].height = 60


# ── Pestaña 2: Contactos ───────────────────────────────────────────────────────
_CON_HEADERS = [
    "ID contacto", "ID empresa", "Empresa", "Nombre", "Cargo",
    "Prioridad cargo", "Email", "Confianza email", "LinkedIn", "Teléfono",
    "Fuente", "Email disponible", "Acción sugerida",
]
_CON_WIDTHS = [12, 8, 30, 25, 30, 14, 30, 16, 35, 18, 20, 16, 30]


def _accion_sugerida(contact: dict) -> str:
    if contact.get("email"):
        return "Enviar email frío"
    if contact.get("linkedin"):
        return "Conectar en LinkedIn"
    if contact.get("telefono"):
        return "Llamada en frío"
    return "Buscar contacto"


def _build_contactos(ws, companies: list[dict]) -> None:
    _write_headers(ws, _CON_HEADERS, _CON_WIDTHS)
    row_i = 2
    for emp in companies:
        eid = emp.get("id_empresa", "")
        nombre_emp = emp.get("nombre_empresa", "")
        for c in emp.get("contactos", []):
            row = [
                c.get("id_contacto"),
                eid,
                nombre_emp,
                c.get("nombre"),
                c.get("cargo"),
                c.get("cargo_prioridad"),
                c.get("email"),
                c.get("confianza_email"),
                c.get("linkedin"),
                c.get("telefono"),
                c.get("fuente"),
                "Sí" if c.get("email") else "No",
                _accion_sugerida(c),
            ]
            for col_i, val in enumerate(row, start=1):
                _set_cell(ws, row_i, col_i, val)
            row_i += 1


# ── Pestaña 3: Mensajes ────────────────────────────────────────────────────────
_MSG_HEADERS = [
    "ID contacto", "Empresa", "Contacto", "Cargo",
    "Canal", "Asunto (email)", "Mensaje", "Estado envío",
]
_MSG_WIDTHS = [12, 30, 25, 30, 12, 40, 70, 16]


def _build_mensajes(ws, mensajes: list[dict]) -> None:
    _write_headers(ws, _MSG_HEADERS, _MSG_WIDTHS)
    row_i = 2
    for m in mensajes:
        cid = m.get("id_contacto", "")
        empresa = m.get("empresa", "")
        contacto = m.get("contacto", "")
        cargo = m.get("cargo", "")
        canales = m.get("mensajes", {})

        if "email" in canales:
            e = canales["email"]
            row = [cid, empresa, contacto, cargo, "Email", e.get("asunto"), e.get("cuerpo"), ""]
            for col_i, val in enumerate(row, start=1):
                _set_cell(ws, row_i, col_i, val, wrap=(col_i == 7))
            ws.row_dimensions[row_i].height = 80
            row_i += 1

        if "linkedin" in canales:
            row = [cid, empresa, contacto, cargo, "LinkedIn", "", canales["linkedin"].get("mensaje"), ""]
            for col_i, val in enumerate(row, start=1):
                _set_cell(ws, row_i, col_i, val, wrap=(col_i == 7))
            ws.row_dimensions[row_i].height = 50
            row_i += 1

        if "whatsapp" in canales:
            row = [cid, empresa, contacto, cargo, "WhatsApp", "", canales["whatsapp"].get("mensaje"), ""]
            for col_i, val in enumerate(row, start=1):
                _set_cell(ws, row_i, col_i, val, wrap=(col_i == 7))
            ws.row_dimensions[row_i].height = 50
            row_i += 1


# ── Pestaña 4: Pipeline ────────────────────────────────────────────────────────
_PIP_HEADERS = [
    "ID empresa", "Empresa", "Score", "Clasificación", "Plan sugerido",
    "Contacto principal", "Cargo contacto", "Canal preferido",
    "Etapa", "Fecha primer contacto", "Próxima acción",
    "Responsable comercial", "Notas",
]
_PIP_WIDTHS = [10, 35, 10, 14, 16, 25, 30, 16, 18, 22, 30, 25, 40]


def _canal_preferido(contacts: list[dict]) -> str:
    for c in contacts:
        if c.get("email"):
            return "Email"
    for c in contacts:
        if c.get("linkedin"):
            return "LinkedIn"
    for c in contacts:
        if c.get("telefono"):
            return "WhatsApp/llamada"
    return ""


def _build_pipeline(ws, companies: list[dict], scores_map: dict) -> None:
    _write_headers(ws, _PIP_HEADERS, _PIP_WIDTHS)
    # Ordenar por score descendente
    sorted_companies = sorted(
        companies,
        key=lambda e: scores_map.get(e.get("id_empresa", ""), {}).get("score") or 0,
        reverse=True,
    )
    for row_i, emp in enumerate(sorted_companies, start=2):
        eid = emp.get("id_empresa", "")
        score_info = scores_map.get(eid, {})
        score = score_info.get("score")
        contacts = emp.get("contactos", [])
        principal = contacts[0] if contacts else {}
        row = [
            eid,
            emp.get("nombre_empresa"),
            score,
            score_info.get("clasificacion"),
            score_info.get("plan_sugerido"),
            principal.get("nombre"),
            principal.get("cargo"),
            _canal_preferido(contacts),
            "",   # Etapa — comercial llena
            "",   # Fecha primer contacto
            "",   # Próxima acción
            "",   # Responsable comercial
            "",   # Notas
        ]
        for col_i, val in enumerate(row, start=1):
            _set_cell(ws, row_i, col_i, val, score=score)


# ── Exportador principal ───────────────────────────────────────────────────────

class CRMExporter:
    """
    Genera el archivo Excel final de prospección.

    Parámetros:
        output_dir: directorio donde guardar el archivo (por defecto /output dentro del contenedor)
    """

    def __init__(self, output_dir: str = "/output") -> None:
        self.output_dir = Path(output_dir)

    def export(
        self,
        companies: list[dict],       # empresas enriquecidas (researcher + enricher)
        scores: list[dict],           # salida del scorer — leads_calificados
        messages: list[dict],         # salida del message_writer — mensajes
    ) -> str:
        """
        Genera el Excel y devuelve la ruta absoluta del archivo creado.

        Solo incluye empresas con incluir_en_pipeline=True.
        """
        self.output_dir.mkdir(parents=True, exist_ok=True)

        # Indexar scores por id_empresa
        scores_map: dict[str, dict] = {s["id_empresa"]: s for s in scores}

        # Filtrar empresas excluidas por el scorer
        companies_in = [
            c for c in companies
            if scores_map.get(c.get("id_empresa", ""), {}).get("incluir_en_pipeline", True)
        ]

        wb = Workbook()
        ws_emp = wb.active
        ws_emp.title = "Empresas"
        ws_con = wb.create_sheet("Contactos")
        ws_msg = wb.create_sheet("Mensajes")
        ws_pip = wb.create_sheet("Pipeline")

        _build_empresas(ws_emp, companies_in, scores_map)
        _build_contactos(ws_con, companies_in)
        _build_mensajes(ws_msg, messages)
        _build_pipeline(ws_pip, companies_in, scores_map)

        # Nombre de archivo con timestamp; si ya existe agrega sufijo _v2, _v3...
        base = f"senal_leads_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        path = self.output_dir / base

        wb.save(str(path))
        return str(path)

    def summary(self, scores: list[dict]) -> dict:
        calientes = sum(1 for s in scores if s.get("clasificacion") == "caliente")
        tibios = sum(1 for s in scores if s.get("clasificacion") == "tibio")
        frios = sum(1 for s in scores if s.get("clasificacion") == "frío")
        descartados = sum(1 for s in scores if not s.get("incluir_en_pipeline", True))
        return {
            "total": len(scores),
            "calientes": calientes,
            "tibios": tibios,
            "frios": frios,
            "descartados": descartados,
        }
