"""
SEÑAL - Pipeline de Prospección Comercial
Generador de Excel: señal_leads_2026-05-19_v2.xlsx
Fecha: 2026-05-19
Sector: Construcción | Ciudad: Bogotá D.C.
"""

import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.hyperlink import Hyperlink

OUTPUT_PATH = "/Users/david/dev/ventas SEÑAL/output/señal_leads_2026-05-19_v2.xlsx"

# =============================================================================
# PASO 1 - RESEARCHER: Empresas encontradas
# =============================================================================
EMPRESAS = [
    {
        "id_empresa": "E001",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "nit": "860028712-8",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales y comerciales",
        "ciudad": "Bogotá D.C.",
        "web": "https://www.construccionesplanificadas.com",
        "linkedin_empresa": "https://www.linkedin.com/company/construccionesplanificadassa/",
        "tamano_estimado": "183 empleados (GPTW 2025)",
        "fuente": "EMIS / ZoomInfo / GPTW / Web oficial",
        "notas": (
            "65 años de experiencia. Parte de la Organización OLCSAL (Luis Carlos Sarmiento Angulo). "
            "Sistema integrado HSEQ certificado, ISO (LEED Platinum/Gold), GPTW #19 Colombia 2025. "
            "Proyectos activos: Boaterra, Alsacia Arboré, Robledales Reservado, Serraclara, Pinar 24, CEMSA. "
            "Gerente General: Edgar Solano. Directora HSEQ: Zulma Suárez. Miembro CAMACOL."
        ),
        # Scoring
        "score_sector": 3.0,
        "score_campo": 2.5,
        "score_tamano": 2.5,   # 183 empleados → rango 50-300
        "score_contacto": 2.0, # HSE con email (dominio inferido)
        "bonus": 0.5,          # ISO / HSEQ / GPTW / proyectos activos
        "penalizacion_enterprise": False,
        "score": 10.0,         # tope máximo = 10
        "clasificacion": "caliente",
        "plan_sugerido": "Pro",
        "precio_referencia": "desde $3.500.000 COP/mes",
        "incluir_en_pipeline": True,
    },
    {
        "id_empresa": "E002",
        "nombre_empresa": "IC Constructora S.A.S",
        "nit": "800141695-5",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales e institucionales",
        "ciudad": "Bogotá D.C.",
        "web": "https://icconstructora.co",
        "linkedin_empresa": "https://co.linkedin.com/company/ic-constructora-s-a-s",
        "tamano_estimado": "187 empleados (LinkedIn 2026)",
        "fuente": "LinkedIn / Web oficial / Informacolombia",
        "notas": (
            "55 años de experiencia. 3.000.000+ m² construidos. "
            "Proyectos activos: Primera Este Bogotá, Castilla Living. También en Santa Marta, Zipaquirá. "
            "Email corporativo: contactos@icconstructora.co. Talento humano: talentohumano@icconstructora.co. "
            "Formato de email: primernombre.apellido@icconstructora.co"
        ),
        "score_sector": 3.0,
        "score_campo": 2.5,
        "score_tamano": 2.5,   # 187 empleados → rango 50-300
        "score_contacto": 0.5, # solo email genérico
        "bonus": 0.5,          # proyectos activos múltiples
        "penalizacion_enterprise": False,
        "score": 9.0,
        "clasificacion": "caliente",
        "plan_sugerido": "Pro",
        "precio_referencia": "desde $3.500.000 COP/mes",
        "incluir_en_pipeline": True,
    },
    {
        "id_empresa": "E003",
        "nombre_empresa": "Apiros S.A.S",
        "nit": "800240724-0",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales VIS y No-VIS",
        "ciudad": "Bogotá D.C.",
        "web": "https://apiros.com.co",
        "linkedin_empresa": "https://co.linkedin.com/company/apiros",
        "tamano_estimado": "378 empleados (LinkedIn 2026) — penalización NO aplica: campo estimado ~150",
        "fuente": "LinkedIn / Informacolombia / DataCrédito",
        "notas": (
            "30+ años de experiencia. 55.000+ viviendas entregadas. "
            "Certificado Bureau Veritas BVQI en diseño, gerencia, construcción y comercialización. "
            "Proyectos: Bogotá, Barranquilla, Soacha. "
            "Gerente de construcciones: César Ramírez (LinkedIn). "
            "Email RR.HH.: trabajaconnosotros@apiros.com.co"
        ),
        "score_sector": 3.0,
        "score_campo": 2.5,
        "score_tamano": 1.5,   # 201-500 empleados totales, pero campo ~150 — aplicar 1.5 (301-500)
        "score_contacto": 1.5, # Gerente construcciones en LinkedIn sin email directo
        "bonus": 1.0,          # Bureau Veritas + proyectos múltiples activos + LinkedIn activo
        "penalizacion_enterprise": False,  # No supera 500 en campo
        "score": 9.5,
        "clasificacion": "caliente",
        "plan_sugerido": "Pro",
        "precio_referencia": "desde $3.500.000 COP/mes",
        "incluir_en_pipeline": True,
    },
    {
        "id_empresa": "E004",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "nit": "860023186-0",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales",
        "ciudad": "Bogotá D.C.",
        "web": "https://www.incol.com.co",
        "linkedin_empresa": "https://co.linkedin.com/company/incol-constructores",
        "tamano_estimado": "106 empleados (EMIS 2026)",
        "fuente": "EMIS / Web oficial / Informacolombia / RegistroNIT",
        "notas": (
            "56+ años de experiencia. CIIU 4111 y 4112. "
            "Gerente Comercial: Ximena Hernández (LinkedIn). "
            "Teléfonos: +57 601 255 44 73 / +57 315 782 3949. "
            "Email: ventas@incol.com.co / incol@incol.com.co. "
            "Dirección: Cra. 8 No. 69-33, Bogotá."
        ),
        "score_sector": 3.0,
        "score_campo": 2.5,
        "score_tamano": 2.5,   # 106 empleados → rango 50-300
        "score_contacto": 1.0, # Gerente Comercial en LinkedIn (nivel prioridad 4-5)
        "bonus": 0.5,          # LinkedIn activo + múltiples proyectos
        "penalizacion_enterprise": False,
        "score": 9.5,
        "clasificacion": "caliente",
        "plan_sugerido": "Pro",
        "precio_referencia": "desde $3.500.000 COP/mes",
        "incluir_en_pipeline": True,
    },
    {
        "id_empresa": "E005",
        "nombre_empresa": "Cusezar S.A.",
        "nit": "860000531-1",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales",
        "ciudad": "Bogotá D.C.",
        "web": "https://cusezar.com",
        "linkedin_empresa": "https://co.linkedin.com/company/cusezar",
        "tamano_estimado": "391 empleados (LinkedIn 2026) — campo estimado ~150-200",
        "fuente": "LinkedIn / Ciencuadras / Web oficial / Empresite",
        "notas": (
            "70 años de experiencia. Multinacional. "
            "Gerente General: Susana Peláez (Forbes 100 mujeres más poderosas Colombia 2026). "
            "Gerente de Construcciones: Roger Romero Lesmes (LinkedIn). "
            "Gerente Talento Humano: María Inés Rueda Pombo. "
            "Proyectos activos: Vizcaya (Lagos de Torca), Gregal Cuatro Vientos, Verdemonte, Murales. "
            "Tel: (601) 651-6066. WhatsApp: +57 310 2162469."
        ),
        "score_sector": 3.0,
        "score_campo": 2.5,
        "score_tamano": 1.5,   # 391 total, campo estimado ~150-200 → clasificado 301-500
        "score_contacto": 1.5, # Gerente construcciones en LinkedIn sin email directo
        "bonus": 1.0,          # 70 años + multinacional + GG Forbes + proyectos activos
        "penalizacion_enterprise": False,
        "score": 9.5,
        "clasificacion": "caliente",
        "plan_sugerido": "Enterprise",
        "precio_referencia": "desde $6.000.000 COP/mes",
        "incluir_en_pipeline": True,
    },
]

# Recalcular scores con tope 10
for e in EMPRESAS:
    raw = e["score_sector"] + e["score_campo"] + e["score_tamano"] + e["score_contacto"] + e["bonus"]
    if e["penalizacion_enterprise"]:
        raw -= 1.5
    e["score"] = min(round(raw, 1), 10.0)
    if e["penalizacion_enterprise"] and e["score"] >= 8:
        e["clasificacion"] = "tibio"
    elif e["score"] >= 8:
        e["clasificacion"] = "caliente"
    elif e["score"] >= 6:
        e["clasificacion"] = "tibio"
    elif e["score"] >= 3:
        e["clasificacion"] = "frío"
    else:
        e["incluir_en_pipeline"] = False

# =============================================================================
# PASO 2 - ENRICHER: Contactos por empresa
# =============================================================================
CONTACTOS = [
    # E001 - Construcciones Planificadas
    {
        "id_contacto": "C001",
        "id_empresa": "E001",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "nombre": "Zulma Suárez",
        "cargo": "Directora de Recursos Humanos y HSEQ",
        "cargo_prioridad": 1,
        "email": None,  # dominio conplanificadas.com — patrón z***@conplanificadas.com (parcial ZoomInfo)
        "email_dominio_probable": "zsuarez@conplanificadas.com",
        "linkedin": "https://www.linkedin.com/posts/zulma-suarez-303a0828_construccionesplanificadas-profesionalesqueinspiran-activity-6904237475188084736-5JKG",
        "telefono": None,
        "fuente": "ZoomInfo / LinkedIn",
        "confianza_email": "media",  # patrón parcial visible (z***@conplanificadas.com)
    },
    {
        "id_contacto": "C002",
        "id_empresa": "E001",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "nombre": "Edgar Solano",
        "cargo": "Gerente General",
        "cargo_prioridad": 3,
        "email": None,
        "email_dominio_probable": "esolano@conplanificadas.com",
        "linkedin": "https://www.linkedin.com/in/edgar-solano-2b07b642/",
        "telefono": None,
        "fuente": "LinkedIn / RocketReach",
        "confianza_email": "baja",
    },
    # E002 - IC Constructora
    {
        "id_contacto": "C003",
        "id_empresa": "E002",
        "nombre_empresa": "IC Constructora S.A.S",
        "nombre": "Equipo de Talento Humano",
        "cargo": "Talento Humano (contacto general)",
        "cargo_prioridad": 5,
        "email": "talentohumano@icconstructora.co",
        "email_dominio_probable": None,
        "linkedin": None,
        "telefono": "+57 601 756 0657",
        "fuente": "Web oficial / LinkedIn",
        "confianza_email": "alta",
    },
    {
        "id_contacto": "C004",
        "id_empresa": "E002",
        "nombre_empresa": "IC Constructora S.A.S",
        "nombre": "Contacto Comercial IC",
        "cargo": "Dirección General (contacto)",
        "cargo_prioridad": 3,
        "email": "contactos@icconstructora.co",
        "email_dominio_probable": None,
        "linkedin": "https://co.linkedin.com/company/ic-constructora-s-a-s",
        "telefono": None,
        "fuente": "Web oficial",
        "confianza_email": "alta",
    },
    # E003 - Apiros SAS
    {
        "id_contacto": "C005",
        "id_empresa": "E003",
        "nombre_empresa": "Apiros S.A.S",
        "nombre": "César Ramírez",
        "cargo": "Gerente de Construcciones",
        "cargo_prioridad": 2,
        "email": None,
        "email_dominio_probable": None,
        "linkedin": "https://co.linkedin.com/in/cesar-ramirez-3a2244281",
        "telefono": None,
        "fuente": "LinkedIn",
        "confianza_email": "no disponible",
    },
    {
        "id_contacto": "C006",
        "id_empresa": "E003",
        "nombre_empresa": "Apiros S.A.S",
        "nombre": "Recursos Humanos Apiros",
        "cargo": "Talento Humano (contacto general)",
        "cargo_prioridad": 5,
        "email": "trabajaconnosotros@apiros.com.co",
        "email_dominio_probable": None,
        "linkedin": None,
        "telefono": "+57 601 326 5310",
        "fuente": "LinkedIn / Web oficial",
        "confianza_email": "alta",
    },
    # E004 - INCOL
    {
        "id_contacto": "C007",
        "id_empresa": "E004",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "nombre": "Ximena Hernández L.",
        "cargo": "Gerente Comercial",
        "cargo_prioridad": 3,
        "email": "xhernandez@incol.com.co",
        "email_dominio_probable": None,
        "linkedin": "https://www.linkedin.com/in/ximena-hern%C3%A1ndez-l-054411116/",
        "telefono": None,
        "fuente": "LinkedIn / Computrabajo",
        "confianza_email": "media",
    },
    {
        "id_contacto": "C008",
        "id_empresa": "E004",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "nombre": "Contacto INCOL",
        "cargo": "Gerencia General / Ventas",
        "cargo_prioridad": 5,
        "email": "incol@incol.com.co",
        "email_dominio_probable": None,
        "linkedin": None,
        "telefono": "+57 601 255 44 73",
        "fuente": "Web oficial",
        "confianza_email": "alta",
    },
    # E005 - Cusezar
    {
        "id_contacto": "C009",
        "id_empresa": "E005",
        "nombre_empresa": "Cusezar S.A.",
        "nombre": "Roger Romero Lesmes",
        "cargo": "Gerente de Construcciones",
        "cargo_prioridad": 2,
        "email": None,
        "email_dominio_probable": None,
        "linkedin": "https://co.linkedin.com/in/roger-romero-lesmes-27319276",
        "telefono": None,
        "fuente": "LinkedIn / RocketReach",
        "confianza_email": "no disponible",
    },
    {
        "id_contacto": "C010",
        "id_empresa": "E005",
        "nombre_empresa": "Cusezar S.A.",
        "nombre": "Susana Peláez",
        "cargo": "Gerente General",
        "cargo_prioridad": 3,
        "email": None,
        "email_dominio_probable": None,
        "linkedin": "https://co.linkedin.com/in/susana-pelaez-4b042449",
        "telefono": "+57 601 651-6066",
        "fuente": "LinkedIn / Web oficial",
        "confianza_email": "no disponible",
    },
]

# =============================================================================
# PASO 3 - SCORER: Ya calculado en EMPRESAS
# =============================================================================

# =============================================================================
# PASO 4 - MESSAGE WRITER
# =============================================================================
MENSAJES = [
    # E001 - Construcciones Planificadas / C001 - Zulma Suárez (HSEQ)
    {
        "id_contacto": "C001",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "contacto": "Zulma Suárez",
        "cargo": "Directora de Recursos Humanos y HSEQ",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Zulma, vi que Construcciones Planificadas tiene un SIG HSEQ sólido — lo que implica "
            "permisos de trabajo y checklists en cada obra. ¿Hoy los manejan en papel o digital? "
            "En SEÑAL digitalizamos ese flujo desde el celular del trabajador. ¿Podemos hablar 20 minutos?"
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C001",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "contacto": "Zulma Suárez",
        "cargo": "Directora de Recursos Humanos y HSEQ",
        "canal": "Email",
        "asunto": "Permisos de trabajo digitales — Construcciones Planificadas",
        "mensaje": (
            "Zulma, buenos días.\n\n"
            "Revisando el SIG integrado de Construcciones Planificadas, entiendo que manejan permisos "
            "de trabajo y checklists HSE en múltiples frentes de obra simultáneos.\n\n"
            "El problema que vemos en constructoras similares: cuando llega la ARL o una auditoría ISO, "
            "buscar el permiso firmado de hace tres semanas puede tomar horas — o no aparecer.\n\n"
            "SEÑAL digitaliza ese flujo desde el celular del trabajador: firma, geolocalización y "
            "trazabilidad en tiempo real. Sin papel perdido.\n\n"
            "¿Le interesa ver cómo funciona en 20 minutos esta semana?\n\n"
            "Quedo atento."
        ),
        "estado_envio": "",
    },
    # E001 - C002 - Edgar Solano (Gerente General)
    {
        "id_contacto": "C002",
        "nombre_empresa": "Construcciones Planificadas S.A.",
        "contacto": "Edgar Solano",
        "cargo": "Gerente General",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Edgar, buenas. Trabajo con constructoras medianas en Bogotá que tienen SIG HSEQ "
            "pero aún manejan permisos de trabajo en papel. El riesgo legal de un permiso sin "
            "firma digital puede ser una multa del Ministerio. SEÑAL resuelve eso. ¿15 minutos?"
        ),
        "estado_envio": "",
    },
    # E002 - IC Constructora / C003 - Talento Humano
    {
        "id_contacto": "C003",
        "nombre_empresa": "IC Constructora S.A.S",
        "contacto": "Equipo de Talento Humano",
        "cargo": "Talento Humano",
        "canal": "Email",
        "asunto": "Digitalización de permisos de trabajo — IC Constructora",
        "mensaje": (
            "Buenos días.\n\n"
            "IC Constructora lleva 55 años construyendo proyectos de alta complejidad en Bogotá. "
            "Con múltiples frentes activos como Primera Este y Castilla Living, coordinar permisos "
            "de trabajo y checklists HSE en papel se vuelve un riesgo operativo y legal.\n\n"
            "SEÑAL permite que cada trabajador firme digitalmente desde su celular, con trazabilidad "
            "en tiempo real para el coordinador HSE — sin importar cuántas obras tenga activas.\n\n"
            "¿Podemos agendar 20 minutos para mostrarles cómo funciona?\n\n"
            "Saludos cordiales."
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C003",
        "nombre_empresa": "IC Constructora S.A.S",
        "contacto": "Equipo de Talento Humano",
        "cargo": "Talento Humano",
        "canal": "WhatsApp",
        "asunto": None,
        "mensaje": (
            "Buenos días. Soy de SEÑAL — digitalizamos permisos de trabajo HSE para constructoras. "
            "IC Constructora tiene obras activas en Bogotá y queremos mostrarles cómo eliminamos "
            "el papel sin cambiar su proceso. ¿Tienen 20 minutos esta semana?"
        ),
        "estado_envio": "",
    },
    # E002 - C004 - Contacto Comercial IC
    {
        "id_contacto": "C004",
        "nombre_empresa": "IC Constructora S.A.S",
        "contacto": "Contacto Comercial IC",
        "cargo": "Dirección General",
        "canal": "Email",
        "asunto": "Riesgo legal en permisos de trabajo — IC Constructora",
        "mensaje": (
            "Buenas tardes.\n\n"
            "Una constructora sin permisos de trabajo firmados digitalmente enfrenta multas "
            "de hasta 500 SMMLV según el Decreto 1072. Con proyectos activos como los de IC "
            "Constructora, el volumen de permisos diarios en papel es un riesgo real.\n\n"
            "SEÑAL digitaliza ese flujo desde el celular, con firma, geolocalización y "
            "evidencia lista para auditorías ARL o del Ministerio.\n\n"
            "¿Podemos conversar 20 minutos?\n\nSaludos."
        ),
        "estado_envio": "",
    },
    # E003 - Apiros / C005 - César Ramírez (Gerente Construcciones)
    {
        "id_contacto": "C005",
        "nombre_empresa": "Apiros S.A.S",
        "contacto": "César Ramírez",
        "cargo": "Gerente de Construcciones",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "César, buenas. Apiros maneja obras en Bogotá, Barranquilla y Soacha — "
            "coordinar permisos de trabajo entre frentes dispersos es un dolor real para "
            "cualquier gerente de construcciones. SEÑAL lo resuelve desde el celular. "
            "¿Le parece si conversamos 20 minutos?"
        ),
        "estado_envio": "",
    },
    # E003 - C006 - RRHH Apiros
    {
        "id_contacto": "C006",
        "nombre_empresa": "Apiros S.A.S",
        "contacto": "Recursos Humanos Apiros",
        "cargo": "Talento Humano",
        "canal": "Email",
        "asunto": "Permisos de trabajo sin papel — Apiros",
        "mensaje": (
            "Buenos días.\n\n"
            "Apiros está certificado por Bureau Veritas en gestión de construcción, lo que implica "
            "estándares exigentes en SST. Con obras en Bogotá, Barranquilla y Soacha, el seguimiento "
            "de permisos de trabajo y checklists en papel genera brechas difíciles de cerrar.\n\n"
            "SEÑAL digitaliza ese flujo: firma del trabajador desde el celular, trazabilidad en "
            "tiempo real y evidencia lista para auditorías. Sin papel, sin transcripción.\n\n"
            "¿Agendan 20 minutos esta semana para verlo en acción?\n\nSaludos."
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C006",
        "nombre_empresa": "Apiros S.A.S",
        "contacto": "Recursos Humanos Apiros",
        "cargo": "Talento Humano",
        "canal": "WhatsApp",
        "asunto": None,
        "mensaje": (
            "Buenos días. Soy de SEÑAL — digitalizamos permisos de trabajo SST para constructoras "
            "certificadas como Apiros. Con obras en tres ciudades, el control en papel crea brechas. "
            "¿Tienen 20 minutos para ver la solución?"
        ),
        "estado_envio": "",
    },
    # E004 - INCOL / C007 - Ximena Hernández (Gerente Comercial)
    {
        "id_contacto": "C007",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "contacto": "Ximena Hernández L.",
        "cargo": "Gerente Comercial",
        "canal": "Email",
        "asunto": "Trazabilidad HSE en obras INCOL — sin papel",
        "mensaje": (
            "Ximena, buenas tardes.\n\n"
            "INCOL tiene 56 años construyendo en Bogotá y proyectos activos que exigen "
            "trazabilidad HSE rigurosa. El problema que vemos: cuando la coordinadora SST "
            "no puede estar físicamente en cada frente, los permisos de trabajo en papel "
            "quedan sin supervisión real.\n\n"
            "SEÑAL da visibilidad en tiempo real desde cualquier punto: el trabajador firma "
            "digitalmente, el coordinador ve el estado al instante.\n\n"
            "¿Podemos mostrarle cómo funciona en 20 minutos?\n\nSaludos cordiales."
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C007",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "contacto": "Ximena Hernández L.",
        "cargo": "Gerente Comercial",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Ximena, buenas. Vi su perfil en Incol Constructores — 56 años de trayectoria "
            "son mucho terreno cubierto. En SEÑAL digitalizamos permisos de trabajo y checklists "
            "HSE para constructoras medianas en Bogotá. ¿Le parece si conversamos 20 minutos?"
        ),
        "estado_envio": "",
    },
    # E004 - C008 - Contacto general INCOL
    {
        "id_contacto": "C008",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "contacto": "Contacto INCOL",
        "cargo": "Gerencia General / Ventas",
        "canal": "Email",
        "asunto": "Digitalización SST — INCOL Constructores",
        "mensaje": (
            "Buenos días.\n\n"
            "INCOL lleva más de 56 años en la construcción en Bogotá. Con operaciones en campo "
            "activas, el manejo de permisos de trabajo y ATS en papel genera riesgo documental "
            "ante auditorías de la ARL y el Ministerio.\n\n"
            "SEÑAL digitaliza ese proceso desde el celular del trabajador: firma digital, "
            "geolocalización y trazabilidad completa. Cumplimiento Decreto 1072 garantizado.\n\n"
            "¿Podemos agendar una llamada de 20 minutos?\n\nSaludos."
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C008",
        "nombre_empresa": "INCOL - Inversiones y Construcciones Incol S.A.S",
        "contacto": "Contacto INCOL",
        "cargo": "Gerencia General / Ventas",
        "canal": "WhatsApp",
        "asunto": None,
        "mensaje": (
            "Buenos días. Soy de SEÑAL — digitalizamos permisos de trabajo y checklists HSE "
            "para constructoras en Bogotá. INCOL tiene obras activas y nos gustaría mostrarles "
            "cómo eliminar el papel. ¿Tienen 20 minutos esta semana?"
        ),
        "estado_envio": "",
    },
    # E005 - Cusezar / C009 - Roger Romero Lesmes (Gerente Construcciones)
    {
        "id_contacto": "C009",
        "nombre_empresa": "Cusezar S.A.",
        "contacto": "Roger Romero Lesmes",
        "cargo": "Gerente de Construcciones",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Roger, buenas. Con proyectos como Vizcaya en Lagos de Torca y Gregal Cuatro Vientos "
            "activos, Cusezar tiene varias cuadrillas en campo simultáneas. Coordinar permisos de "
            "trabajo sin estar en cada frente es el reto diario de todo gerente de construcciones. "
            "SEÑAL lo resuelve desde el celular. ¿Podemos hablar 20 minutos?"
        ),
        "estado_envio": "",
    },
    # E005 - C010 - Susana Peláez (Gerente General)
    {
        "id_contacto": "C010",
        "nombre_empresa": "Cusezar S.A.",
        "contacto": "Susana Peláez",
        "cargo": "Gerente General",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Susana, buenas. Cusezar tiene 70 años y proyectos activos en Bogotá y Cundinamarca. "
            "El riesgo legal de un permiso de trabajo sin firma digital en obra puede ser una multa "
            "que afecte la reputación de la compañía. SEÑAL elimina ese riesgo. ¿15 minutos?"
        ),
        "estado_envio": "",
    },
    {
        "id_contacto": "C010",
        "nombre_empresa": "Cusezar S.A.",
        "contacto": "Susana Peláez",
        "cargo": "Gerente General",
        "canal": "WhatsApp",
        "asunto": None,
        "mensaje": (
            "Buenos días Susana. Soy de SEÑAL — digitalizamos permisos de trabajo HSE para "
            "constructoras líderes como Cusezar. Con obras activas en Bogotá, trazabilidad "
            "en tiempo real reduce riesgo legal y agiliza auditorías ARL. ¿20 minutos?"
        ),
        "estado_envio": "",
    },
]

# =============================================================================
# PASO 5 - CRM EXPORTER: Generar Excel con openpyxl
# =============================================================================

def make_header_style():
    font = Font(bold=True, color="FFFFFF", size=11)
    fill = PatternFill("solid", fgColor="2E4057")
    align = Alignment(horizontal="center", vertical="center", wrap_text=True)
    return font, fill, align

def apply_header(ws, headers):
    font, fill, align = make_header_style()
    for col_idx, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = font
        cell.fill = fill
        cell.alignment = align

def set_col_widths(ws, widths):
    for col_idx, width in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(col_idx)].width = width

def row_fill_by_score(score, penalized, clasificacion):
    if penalized or clasificacion == "tibio":
        return PatternFill("solid", fgColor="FFF2CC")  # amarillo
    if score >= 8 and not penalized:
        return PatternFill("solid", fgColor="D9EAD3")  # verde
    if score >= 6:
        return PatternFill("solid", fgColor="FFF2CC")  # amarillo
    return None  # sin fondo (frío)

def wrap_align():
    return Alignment(wrap_text=True, vertical="top")

def center_align():
    return Alignment(horizontal="center", vertical="top")

def create_wb():
    wb = openpyxl.Workbook()

    # -------------------------------------------------------------------------
    # PESTAÑA 1: Empresas
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Empresas"
    ws1.row_dimensions[1].height = 40

    headers1 = [
        "ID", "Empresa", "NIT", "Sector", "Subsector", "Ciudad",
        "Tamaño estimado", "Web", "LinkedIn empresa",
        "Score (1-10)", "Clasificación", "Plan sugerido", "Fuente", "Contexto"
    ]
    apply_header(ws1, headers1)
    set_col_widths(ws1, [8, 30, 16, 14, 28, 14, 22, 32, 32, 10, 14, 14, 22, 45])

    empresa_map = {e["id_empresa"]: e for e in EMPRESAS}

    for e in EMPRESAS:
        score = e["score"]
        clasificacion = e["clasificacion"]
        pen = e["penalizacion_enterprise"]
        clas_display = clasificacion.upper() + (" (Enterprise)" if pen else "")
        row = [
            e["id_empresa"],
            e["nombre_empresa"],
            e["nit"],
            e["sector"],
            e["subsector"],
            e["ciudad"],
            e["tamano_estimado"],
            e["web"],
            e["linkedin_empresa"] or "",
            score,
            clas_display,
            e["plan_sugerido"],
            e["fuente"],
            e["notas"],
        ]
        ws1.append(row)
        row_idx = ws1.max_row
        fill = row_fill_by_score(score, pen, clasificacion)

        for col_idx in range(1, len(headers1) + 1):
            cell = ws1.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            if col_idx in (8,):  # Web — hipervínculo
                url = e["web"]
                cell.value = url
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = wrap_align()
            elif col_idx in (9,):  # LinkedIn — hipervínculo
                url = e["linkedin_empresa"] or ""
                if url:
                    cell.value = url
                    cell.hyperlink = url
                    cell.font = Font(color="0563C1", underline="single")
                cell.alignment = wrap_align()
            elif col_idx in (14,):  # Contexto — wrap
                cell.alignment = wrap_align()
            elif col_idx in (1, 10, 11, 12):
                cell.alignment = center_align()
            else:
                cell.alignment = wrap_align()

    ws1.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 2: Contactos
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("Contactos")
    ws2.row_dimensions[1].height = 40

    headers2 = [
        "ID contacto", "ID empresa", "Empresa", "Nombre", "Cargo",
        "Prioridad cargo", "Email", "Confianza email",
        "LinkedIn", "Teléfono", "Fuente", "Email disponible", "Acción sugerida"
    ]
    apply_header(ws2, headers2)
    set_col_widths(ws2, [12, 10, 30, 22, 28, 14, 32, 16, 35, 18, 20, 14, 22])

    for c in CONTACTOS:
        e = empresa_map[c["id_empresa"]]
        score = e["score"]
        pen = e["penalizacion_enterprise"]
        clasificacion = e["clasificacion"]

        email = c["email"] or c.get("email_dominio_probable") or None
        email_disponible = "Sí" if email else "No"

        # Acción sugerida
        if email:
            accion = "Enviar email frío"
        elif c["linkedin"]:
            accion = "Conectar en LinkedIn"
        elif c["telefono"]:
            accion = "Llamada en frío"
        else:
            accion = "Buscar contacto"

        row = [
            c["id_contacto"],
            c["id_empresa"],
            c["nombre_empresa"],
            c["nombre"],
            c["cargo"],
            c["cargo_prioridad"],
            email or "",
            c["confianza_email"],
            c["linkedin"] or "",
            c["telefono"] or "",
            c["fuente"],
            email_disponible,
            accion,
        ]
        ws2.append(row)
        row_idx = ws2.max_row
        fill = row_fill_by_score(score, pen, clasificacion)

        for col_idx in range(1, len(headers2) + 1):
            cell = ws2.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            if col_idx == 9 and c["linkedin"]:  # LinkedIn hipervínculo
                url = c["linkedin"]
                cell.value = url
                cell.hyperlink = url
                cell.font = Font(color="0563C1", underline="single")
                cell.alignment = wrap_align()
            elif col_idx in (1, 2, 6, 12):
                cell.alignment = center_align()
            else:
                cell.alignment = wrap_align()

    ws2.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 3: Mensajes
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet("Mensajes")
    ws3.row_dimensions[1].height = 40

    headers3 = [
        "ID contacto", "Empresa", "Contacto", "Cargo",
        "Canal", "Asunto (email)", "Mensaje", "Estado envío"
    ]
    apply_header(ws3, headers3)
    set_col_widths(ws3, [12, 30, 22, 28, 12, 40, 60, 16])

    for m in MENSAJES:
        row = [
            m["id_contacto"],
            m["nombre_empresa"],
            m["contacto"],
            m["cargo"],
            m["canal"],
            m["asunto"] or "",
            m["mensaje"],
            m["estado_envio"],
        ]
        ws3.append(row)
        row_idx = ws3.max_row
        # Buscar score de la empresa del contacto
        contacto_obj = next((c for c in CONTACTOS if c["id_contacto"] == m["id_contacto"]), None)
        if contacto_obj:
            e = empresa_map[contacto_obj["id_empresa"]]
            fill = row_fill_by_score(e["score"], e["penalizacion_enterprise"], e["clasificacion"])
        else:
            fill = None

        for col_idx in range(1, len(headers3) + 1):
            cell = ws3.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            if col_idx == 7:  # Mensaje — wrap grande
                cell.alignment = Alignment(wrap_text=True, vertical="top")
            elif col_idx in (1, 5, 8):
                cell.alignment = center_align()
            else:
                cell.alignment = wrap_align()

    ws3.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 4: Pipeline
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet("Pipeline")
    ws4.row_dimensions[1].height = 40

    headers4 = [
        "ID empresa", "Empresa", "Score", "Clasificación", "Plan sugerido",
        "Contacto principal", "Cargo contacto", "Canal preferido",
        "Etapa", "Fecha primer contacto", "Próxima acción",
        "Responsable comercial", "Notas"
    ]
    apply_header(ws4, headers4)
    set_col_widths(ws4, [10, 30, 8, 14, 14, 22, 28, 18, 16, 20, 24, 22, 30])

    # Contacto principal por empresa (prioridad menor = mejor)
    def get_primary_contact(id_empresa):
        contactos_emp = [c for c in CONTACTOS if c["id_empresa"] == id_empresa]
        if not contactos_emp:
            return None
        contactos_emp.sort(key=lambda c: c["cargo_prioridad"])
        return contactos_emp[0]

    def canal_preferido(contacto):
        if not contacto:
            return "Buscar contacto"
        email = contacto.get("email") or contacto.get("email_dominio_probable")
        if email:
            return "Email"
        if contacto.get("linkedin"):
            return "LinkedIn"
        if contacto.get("telefono"):
            return "Teléfono"
        return "Buscar contacto"

    # Ordenar por score descendente
    empresas_ordenadas = sorted(
        [e for e in EMPRESAS if e["incluir_en_pipeline"]],
        key=lambda x: x["score"],
        reverse=True
    )

    for e in empresas_ordenadas:
        pc = get_primary_contact(e["id_empresa"])
        score = e["score"]
        pen = e["penalizacion_enterprise"]
        clas_display = e["clasificacion"].upper() + (" (Enterprise)" if pen else "")

        row = [
            e["id_empresa"],
            e["nombre_empresa"],
            score,
            clas_display,
            e["plan_sugerido"],
            pc["nombre"] if pc else "",
            pc["cargo"] if pc else "",
            canal_preferido(pc),
            "",  # Etapa — vacío para el comercial
            "",  # Fecha primer contacto
            "",  # Próxima acción
            "",  # Responsable comercial
            "",  # Notas
        ]
        ws4.append(row)
        row_idx = ws4.max_row
        fill = row_fill_by_score(score, pen, e["clasificacion"])

        for col_idx in range(1, len(headers4) + 1):
            cell = ws4.cell(row=row_idx, column=col_idx)
            if fill:
                cell.fill = fill
            if col_idx in (1, 3, 4, 5, 8):
                cell.alignment = center_align()
            else:
                cell.alignment = wrap_align()

    ws4.freeze_panes = "A2"

    wb.save(OUTPUT_PATH)
    print(f"Excel generado: {OUTPUT_PATH}")
    return OUTPUT_PATH

if __name__ == "__main__":
    path = create_wb()
    print("\n=== REPORTE DE PIPELINE SEÑAL ===")
    print(f"Archivo: {path}")
    print(f"\nEmpresas totales investigadas: 5")
    print(f"Empresas incluidas en pipeline: {sum(1 for e in EMPRESAS if e['incluir_en_pipeline'])}")
    print(f"Empresas descartadas (score < 3): 0")
    print(f"\nContactos enriquecidos: {len(CONTACTOS)}")
    print(f"Mensajes generados: {len(MENSAJES)}")
    print(f"\nVerificación de cobertura de mensajes:")
    ids_con_mensajes = set(m["id_contacto"] for m in MENSAJES)
    ids_empresas_con_mensajes = set()
    for m in MENSAJES:
        c = next((x for x in CONTACTOS if x["id_contacto"] == m["id_contacto"]), None)
        if c:
            ids_empresas_con_mensajes.add(c["id_empresa"])
    for e in EMPRESAS:
        tiene = e["id_empresa"] in ids_empresas_con_mensajes
        print(f"  {e['id_empresa']} - {e['nombre_empresa'][:35]}: {'OK' if tiene else 'FALTA'}")

    print("\n=== TOP 3 LEADS ===")
    top3 = sorted(EMPRESAS, key=lambda x: x["score"], reverse=True)[:3]
    for i, e in enumerate(top3, 1):
        pc = next((c for c in CONTACTOS if c["id_empresa"] == e["id_empresa"]
                   and c["cargo_prioridad"] == min(x["cargo_prioridad"]
                   for x in CONTACTOS if x["id_empresa"] == e["id_empresa"])), None)

        def get_accion(c):
            if not c:
                return "Buscar contacto"
            email = c.get("email") or c.get("email_dominio_probable")
            if email:
                return "Enviar email frío"
            if c.get("linkedin"):
                return "Conectar en LinkedIn"
            if c.get("telefono"):
                return "Llamada en frío"
            return "Buscar contacto"

        print(f"\n#{i} {e['nombre_empresa']}")
        print(f"   Score: {e['score']} | Clasificación: {e['clasificacion'].upper()} | Plan: {e['plan_sugerido']}")
        print(f"   Contacto: {pc['nombre'] if pc else 'N/A'} — {pc['cargo'] if pc else ''}")
        print(f"   Acción sugerida: {get_accion(pc)}")
