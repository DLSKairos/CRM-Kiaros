#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Pipeline de prospección comercial SEÑAL
Genera señal_leads_2026-05-19.xlsx con datos reales de 5 constructoras en Bogotá
"""

import os
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

OUTPUT_PATH = "/Users/david/dev/ventas SEÑAL/output/señal_leads_2026-05-19.xlsx"

# =============================================================================
# PASO 1 + 2: DATOS DE EMPRESAS Y CONTACTOS (Researcher + Enricher)
# =============================================================================

EMPRESAS = [
    {
        "id": "E001",
        "nombre": "Constructora Colpatria S.A.S.",
        "nit": "860.058.070-6",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales e infraestructura",
        "ciudad": "Bogotá D.C.",
        "web": "https://www.constructoracolpatria.com",
        "linkedin": "https://www.linkedin.com/company/constructora-colpatria",
        "tamano": "200-500 empleados campo (12 proyectos activos Bogotá)",
        "fuente": "InformaColombia + web oficial + LinkedIn",
        "notas": "NIT 860058070; fundada 1977; 12 proyectos activos en Bogotá; divisiones residencial, comercial, infraestructura; referencia ISO; Andrés Largacha CEO; vacante activa Director SST detectada en mayo 2026.",
        # Scoring
        "score": 8.5,
        "clasificacion": "caliente",
        "plan": "Pro",
        "justificacion": "Gran constructora con 12 obras activas en Bogotá, área SST en búsqueda de director (vacante abierta), confirma necesidad interna de gestión documental en campo. Tamaño y sector ideales para SEÑAL.",
        # Enriquecimiento
        "email_generico": "contactenos@constructoracolpatria.com",
        "telefono": "+57 (601) 744 6050",
        "contexto": "Vacante abierta para Director SST y MA en Bogotá (mayo 2026). 12 proyectos residenciales activos en Bogotá. CEO Andrés Largacha. Gerente Técnico e Innovación: Catalina García Garzón. Director Técnico Infraestructura: David Eduardo González. Presencia en Medellín, Cali, Barranquilla. Referencia a ISO en sitio web.",
    },
    {
        "id": "E002",
        "nombre": "Amarilo S.A.S.",
        "nit": "830.099.399-1",
        "sector": "Construcción",
        "subsector": "Urbanismo y vivienda VIS/No VIS",
        "ciudad": "Bogotá D.C.",
        "web": "https://amarilo.com.co",
        "linkedin": "https://co.linkedin.com/company/constructora-amarilo-sa",
        "tamano": "500+ empleados (13.000 empleos generados, 150+ proyectos activos, 18 ciudades)",
        "fuente": "Wikipedia + web oficial + LinkedIn + Camacol BIM",
        "notas": "Constructora más grande de Colombia según Sectorial/Valoraanalitik 2024. Roberto Moreno presidente. Andrés Pacheco gerente general. 24 nuevos proyectos lanzados 2026. Miembro Camacol. Participación BIM Forum Camacol.",
        # Scoring
        "score": 9.0,
        "clasificacion": "caliente",
        "plan": "Enterprise",
        "justificacion": "La constructora más grande de Colombia con 150+ proyectos activos y 13.000 empleos generados; escala masiva de operaciones en campo en Bogotá y 17 ciudades más. Potencial de rollout nacional desde Bogotá.",
        # Enriquecimiento
        "email_generico": None,
        "telefono": None,
        "contexto": "Líder del sector constructor colombiano. Roberto Moreno (Presidente Ejecutivo), Andrés Pacheco (Gerente General), Omar Ricardo Barbosa Vargas (Director de Proyectos), Andrea Arizala (VP Talento Humano y Comunidades). 24 nuevos proyectos 2026. Miembro activo Camacol y BIM Forum. Presencia LinkedIn activa.",
    },
    {
        "id": "E003",
        "nombre": "Constructora Bolívar S.A.",
        "nit": "860.513.493-1",
        "sector": "Construcción",
        "subsector": "Edificaciones residenciales e infraestructura",
        "ciudad": "Bogotá D.C.",
        "web": "https://www.constructorabolivar.com",
        "linkedin": "https://www.linkedin.com/company/constructora-bolivar-s.a.",
        "tamano": "4.000 empleados (administrativo + campo + ventas + servicio)",
        "fuente": "Web oficial + LinkedIn + BNamericas + InformaColombia",
        "notas": "NIT 860513493-1; fundada 1982, parte del Grupo Bolívar (Seguros Bolívar). 4.000 empleados. Operaciones en 11 ciudades. Bogotá principal sede. Certificaciones LEED y EDGE. Diego Ospina: Director Ejecutivo Nacional de Operaciones. Diego Avendaño: Gerente Técnico y de Construcciones.",
        # Scoring
        "score": 8.0,
        "clasificacion": "caliente",
        "plan": "Enterprise",
        "justificacion": "4.000 empleados con obras residenciales en 11 ciudades y sede principal en Bogotá. Director de Operaciones identificado en LinkedIn con nombre real. Pertenece a Grupo Bolívar, lo que implica estándares corporativos de cumplimiento SST.",
        # Enriquecimiento
        "email_generico": None,
        "telefono": "(+57) 310 315 7550",
        "contexto": "Grupo Bolívar: Seguros Bolívar + Banco Davivienda + Constructora Bolívar. Certificaciones LEED y EDGE. Uno de los 50 mejores empleadores según Merco. Proyectos activos en Soacha (Caminos del Vínculo) y modalidad llave en mano. Diego Ospina identificado como Director Ejecutivo Nacional de Operaciones.",
    },
    {
        "id": "E004",
        "nombre": "Pavimentos Colombia S.A.S.",
        "nit": "860.024.586-8",
        "sector": "Construcción",
        "subsector": "Infraestructura vial y obras civiles",
        "ciudad": "Bogotá D.C.",
        "web": "https://www.pavimentoscolombia.com",
        "linkedin": "https://co.linkedin.com/company/pavimentos-colombia",
        "tamano": "200-500 empleados (55+ años, ventas >100 mil millones COP 2024)",
        "fuente": "InformaColombia + web oficial + RocketReach + LinkedIn",
        "notas": "NIT 8600245868; 55+ años operando; ISO 9001; ONAC LAB-008; Carbono Neutral ICONTEC; ESG Platinum; Transmilenio Zero Waste Gold. Sede: Calle 84A 10-50 Piso 9 Bogotá. Plantas en Sibaté y Sonso. Roberto Herrera VP Comercial, Rodrigo Soto Director Proyecto, Arnoldo Mestre Director Sostenibilidad.",
        # Scoring
        "score": 8.0,
        "clasificacion": "caliente",
        "plan": "Pro",
        "justificacion": "55 años de operaciones viales en campo con certificaciones ISO 9001 y ONAC, lo que implica auditorías periódicas y documentación estricta. Plantas operativas en Sibaté y Cundinamarca. Alto volumen de trabajadores en campo.",
        # Enriquecimiento
        "email_generico": "pavimentos@pavcol.com",
        "telefono": "+57 (601) 376 0030",
        "contexto": "ISO 9001 certificado. Acreditación ONAC LAB-008. Certificación Carbon Neutral ICONTEC. ESG Platinum. Premio Cero Residuos Categoría Oro Proyecto Transmilenio. 55+ años. Ventas >100 mil millones COP (2024). Patrimonio neto 527 mil millones COP. Directivos identificados: Roberto Herrera (VP Comercial), Rodrigo Soto (Director Proyecto), Arnoldo Mestre Trujillo (Director Sostenibilidad).",
    },
    {
        "id": "E005",
        "nombre": "Sacyr Construcción Colombia S.A.S.",
        "nit": "900.657.570-1",
        "sector": "Construcción",
        "subsector": "Infraestructura vial y concesiones",
        "ciudad": "Bogotá D.C.",
        "web": "https://sacyrinfraestructuras.com/en/colombia",
        "linkedin": "https://co.linkedin.com/company/sacyr-colombia",
        "tamano": "348 empleados directos (13.000 empleos en zonas de influencia)",
        "fuente": "InformaColombia + web Sacyr + LinkedIn + EMIS",
        "notas": "NIT 9006575701; filial colombiana de Sacyr S.A. (España); Calle 99 14-49 Piso 4 Torre EAR Bogotá; tel 6017442320; 10+ proyectos activos en Colombia; 4 proyectos con IDU Bogotá; carreteras 4G Rumichaca-Pasto, Pumarejo, Hisgaura. Mauricio Caicedo identificado como Jefe Seguridad Industrial y SST. Javier Fernández: Project Manager Colombia.",
        # Scoring
        "score": 7.5,
        "clasificacion": "tibio",
        "plan": "Pro",
        "justificacion": "348 empleados directos con 10+ proyectos de infraestructura activos incluyendo 4 con el IDU de Bogotá. Filial española con estándares HSE internacionales; el contacto SST fue identificado en LinkedIn con nombre real.",
        # Enriquecimiento
        "email_generico": "cvillamil@sacyr.com",
        "telefono": "+57 (601) 744 2320",
        "contexto": "Sacyr es multinacional española. En Colombia: autopistas 4G, metro Bogotá área, bocatomas Chivor, puentes Pumarejo e Hisgaura, Túnel Occidente. 4 proyectos con Alcaldía de Bogotá (IDU). Empleados actuales 348 directos. Mauricio Caicedo (Jefe Seguridad Industrial y SST - SACYR), Javier Fernández Caamaño (Project Manager SACYR Colombia).",
    },
]

CONTACTOS = [
    # E001 - Constructora Colpatria
    {
        "id": "C001",
        "id_empresa": "E001",
        "empresa": "Constructora Colpatria S.A.S.",
        "nombre": "Andrés Largacha",
        "cargo": "CEO / Gerente General",
        "prioridad": 3,
        "email": None,
        "confianza_email": None,
        "linkedin": None,
        "telefono": "+57 (601) 744 6050",
        "fuente": "RocketReach + medios especializados",
    },
    {
        "id": "C002",
        "id_empresa": "E001",
        "empresa": "Constructora Colpatria S.A.S.",
        "nombre": "Catalina García Garzón",
        "cargo": "Gerente Técnico e Innovación",
        "prioridad": 4,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://rocketreach.co/catalina-garcia-garzon-email_313469534",
        "telefono": None,
        "fuente": "RocketReach",
    },
    {
        "id": "C003",
        "id_empresa": "E001",
        "empresa": "Constructora Colpatria S.A.S.",
        "nombre": "Adriana Marcela Juyó Gómez",
        "cargo": "Representante / Contacto Corporativo",
        "prioridad": 5,
        "email": "contactenos@constructoracolpatria.com",
        "confianza_email": "alto (email genérico verificado en web)",
        "linkedin": None,
        "telefono": "+57 (601) 744 6050",
        "fuente": "InformaColombia + web oficial",
    },
    # E002 - Amarilo
    {
        "id": "C004",
        "id_empresa": "E002",
        "empresa": "Amarilo S.A.S.",
        "nombre": "Roberto Moreno",
        "cargo": "Presidente Ejecutivo",
        "prioridad": 3,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/robertomorenocolombia/",
        "telefono": None,
        "fuente": "LinkedIn + Valoraanalitik + Cambiocolombia",
    },
    {
        "id": "C005",
        "id_empresa": "E002",
        "empresa": "Amarilo S.A.S.",
        "nombre": "Omar Ricardo Barbosa Vargas",
        "cargo": "Director de Proyectos de Construcción",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/omar-ricardo-barbosa-vargas-5967a1136/",
        "telefono": None,
        "fuente": "LinkedIn",
    },
    # E003 - Constructora Bolívar
    {
        "id": "C006",
        "id_empresa": "E003",
        "empresa": "Constructora Bolívar S.A.",
        "nombre": "Diego Ospina",
        "cargo": "Director Ejecutivo Nacional de Operaciones",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/diego-ospina-9476b249/",
        "telefono": None,
        "fuente": "LinkedIn + RocketReach",
    },
    {
        "id": "C007",
        "id_empresa": "E003",
        "empresa": "Constructora Bolívar S.A.",
        "nombre": "Diego Mauricio Avendaño Ruiz",
        "cargo": "Gerente Técnico y de Construcciones",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/diego-avendano-ruiz-construccion-planeacion-control-presupuestos-procesos-tecnologia/",
        "telefono": None,
        "fuente": "LinkedIn",
    },
    # E004 - Pavimentos Colombia
    {
        "id": "C008",
        "id_empresa": "E004",
        "empresa": "Pavimentos Colombia S.A.S.",
        "nombre": "Roberto Herrera",
        "cargo": "Vicepresidente Comercial",
        "prioridad": 3,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/herreraroberto/",
        "telefono": "+57 (601) 376 0030",
        "fuente": "LinkedIn + RocketReach",
    },
    {
        "id": "C009",
        "id_empresa": "E004",
        "empresa": "Pavimentos Colombia S.A.S.",
        "nombre": "Rodrigo Soto",
        "cargo": "Director de Proyecto",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/rodrigo-soto-10253829/",
        "telefono": None,
        "fuente": "LinkedIn",
    },
    {
        "id": "C010",
        "id_empresa": "E004",
        "empresa": "Pavimentos Colombia S.A.S.",
        "nombre": "Arnoldo Mestre Trujillo",
        "cargo": "Director de Sostenibilidad",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://rocketreach.co/arnoldo-mestre-trujillo-email_258275618",
        "telefono": None,
        "fuente": "RocketReach",
    },
    # E005 - Sacyr
    {
        "id": "C011",
        "id_empresa": "E005",
        "empresa": "Sacyr Construcción Colombia S.A.S.",
        "nombre": "Mauricio Caicedo",
        "cargo": "Jefe de Seguridad Industrial y Salud en el Trabajo",
        "prioridad": 1,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/mauricio-caicedo-42316273/",
        "telefono": None,
        "fuente": "LinkedIn",
    },
    {
        "id": "C012",
        "id_empresa": "E005",
        "empresa": "Sacyr Construcción Colombia S.A.S.",
        "nombre": "Javier Fernández Caamaño",
        "cargo": "Project Manager",
        "prioridad": 2,
        "email": None,
        "confianza_email": None,
        "linkedin": "https://www.linkedin.com/in/javier-fern%C3%A1ndez-caama%C3%B1o-976a51131/",
        "telefono": None,
        "fuente": "LinkedIn",
    },
]

# =============================================================================
# PASO 4: MENSAJES PERSONALIZADOS
# =============================================================================

MENSAJES = [
    # E001 - C001 Andrés Largacha (CEO Colpatria)
    {
        "id_contacto": "C001",
        "empresa": "Constructora Colpatria S.A.S.",
        "contacto": "Andrés Largacha",
        "cargo": "CEO / Gerente General",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Andrés, con 12 obras activas en Bogotá y una vacante abierta de Director SST, Colpatria enfrenta el reto de controlar la documentación de campo sin perder trazabilidad. SEÑAL digitaliza permisos de trabajo y checklists HSE desde el celular. ¿Le comparto cómo lo hacemos en 15 minutos?"
        ),
        "estado": "Pendiente",
    },
    # E001 - C002 Catalina García Garzón (Gerente Técnica)
    {
        "id_contacto": "C002",
        "empresa": "Constructora Colpatria S.A.S.",
        "contacto": "Catalina García Garzón",
        "cargo": "Gerente Técnico e Innovación",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Catalina, en constructoras con 12+ proyectos activos como Colpatria, el mayor cuello de botella no es la tecnología sino la adopción en obra. SEÑAL funciona desde el celular del trabajador sin infraestructura adicional. ¿Hablamos esta semana sobre cómo se adapta a su operación?"
        ),
        "estado": "Pendiente",
    },
    # E001 - C003 Adriana Juyó (email genérico)
    {
        "id_contacto": "C003",
        "empresa": "Constructora Colpatria S.A.S.",
        "contacto": "Adriana Marcela Juyó Gómez",
        "cargo": "Representante / Contacto Corporativo",
        "canal": "Email",
        "asunto": "Permisos de trabajo digitales para sus 12 obras en Bogotá",
        "mensaje": (
            "Buenas tardes,\n\n"
            "Con 12 proyectos activos en Bogotá, su equipo de campo gestiona cientos de permisos de trabajo y checklists HSE cada semana. Cuando viene la ARL o la Secretaría del Trabajo, buscar ese papel firmado de hace tres meses puede costarle muy caro.\n\n"
            "SEÑAL digitaliza ese proceso desde el celular: firma en campo, trazabilidad en tiempo real, sin papel perdido. Cumplimiento Decreto 1072 y Resolución 0312.\n\n"
            "¿Le agendamos una demo de 20 minutos esta semana?\n\n"
            "Cordialmente,\nEquipo SEÑAL — Kairos DLS Group S.A.S."
        ),
        "estado": "Pendiente",
    },
    # E002 - C004 Roberto Moreno (Presidente Amarilo)
    {
        "id_contacto": "C004",
        "empresa": "Amarilo S.A.S.",
        "contacto": "Roberto Moreno",
        "cargo": "Presidente Ejecutivo",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Roberto, Amarilo lanza 24 proyectos en 2026. A esa escala, un accidente sin el permiso de trabajo correctamente firmado puede derivar en sanciones que superan por mucho el costo de prevenirlas. SEÑAL da trazabilidad documental completa desde el celular del operario. ¿Le comparto cómo funciona en 15 minutos?"
        ),
        "estado": "Pendiente",
    },
    # E002 - C005 Omar Barbosa (Director Proyectos Amarilo)
    {
        "id_contacto": "C005",
        "empresa": "Amarilo S.A.S.",
        "contacto": "Omar Ricardo Barbosa Vargas",
        "cargo": "Director de Proyectos de Construcción",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Omar, con múltiples frentes de obra activos en Amarilo, ¿cómo garantiza que cada permiso de trabajo está firmado antes de que inicie la actividad crítica? SEÑAL le da ese control en tiempo real sin que tenga que estar físicamente en cada obra. ¿Conversamos esta semana?"
        ),
        "estado": "Pendiente",
    },
    # E003 - C006 Diego Ospina (Director Operaciones Bolívar)
    {
        "id_contacto": "C006",
        "empresa": "Constructora Bolívar S.A.",
        "contacto": "Diego Ospina",
        "cargo": "Director Ejecutivo Nacional de Operaciones",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Diego, con Constructora Bolívar operando en 11 ciudades, controlar que cada frente de obra cumpla con los permisos de trabajo y ATS antes de iniciar actividades críticas es un desafío real. SEÑAL le da visibilidad desde el celular sin estar físicamente en cada obra. ¿Le muestro cómo en 20 minutos?"
        ),
        "estado": "Pendiente",
    },
    # E003 - C007 Diego Avendaño (Gerente Técnico Bolívar)
    {
        "id_contacto": "C007",
        "empresa": "Constructora Bolívar S.A.",
        "contacto": "Diego Mauricio Avendaño Ruiz",
        "cargo": "Gerente Técnico y de Construcciones",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Diego, en Constructora Bolívar con 4.000 empleados y obras en 11 ciudades, el control documental de campo es crítico. SEÑAL digitaliza permisos de trabajo, ATS y checklists HSE con firma en sitio y trazabilidad completa. Sin papel, sin transcripción. ¿Le agendamos una demo?"
        ),
        "estado": "Pendiente",
    },
    # E004 - C008 Roberto Herrera (VP Comercial Pavimentos)
    {
        "id_contacto": "C008",
        "empresa": "Pavimentos Colombia S.A.S.",
        "contacto": "Roberto Herrera",
        "cargo": "Vicepresidente Comercial",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Roberto, Pavimentos Colombia tiene 55 años de operaciones y certificaciones ISO 9001 y ONAC. En el siguiente ciclo de auditoría HSEQ, la trazabilidad de permisos de trabajo en campo puede ser un punto de observación. SEÑAL asegura ese registro digital desde el celular del operario. ¿Le comparto cómo funciona?"
        ),
        "estado": "Pendiente",
    },
    # E004 - C009 Rodrigo Soto (Director Proyecto)
    {
        "id_contacto": "C009",
        "empresa": "Pavimentos Colombia S.A.S.",
        "contacto": "Rodrigo Soto",
        "cargo": "Director de Proyecto",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Rodrigo, en proyectos viales de la escala de Pavimentos Colombia, coordinar los permisos de trabajo de altura y espacios confinados con cuadrillas distribuidas en campo es complejo. SEÑAL digitaliza ese flujo y le da trazabilidad en tiempo real. ¿Conversamos 20 minutos esta semana?"
        ),
        "estado": "Pendiente",
    },
    # E004 - C010 Email genérico Pavimentos
    {
        "id_contacto": "C010",
        "empresa": "Pavimentos Colombia S.A.S.",
        "contacto": "Arnoldo Mestre Trujillo",
        "cargo": "Director de Sostenibilidad",
        "canal": "Email",
        "asunto": "Trazabilidad documental SST para sus proyectos viales",
        "mensaje": (
            "Buenos días,\n\n"
            "Con 55 años de operaciones y certificación ISO 9001, Pavimentos Colombia tiene estándares de calidad que también aplican a la gestión documental SST en campo.\n\n"
            "Cuando llega una auditoría externa o un accidente en obra, encontrar el permiso de trabajo firmado del día correcto puede ser la diferencia entre una observación y una sanción.\n\n"
            "SEÑAL digitaliza permisos de trabajo, ATS y checklists desde el celular, con firma en sitio y trazabilidad lista para auditoría.\n\n"
            "¿Le mostramos cómo funciona en 20 minutos?\n\n"
            "Cordialmente,\nEquipo SEÑAL — Kairos DLS Group S.A.S.\npavimentos@pavcol.com"
        ),
        "estado": "Pendiente",
    },
    # E005 - C011 Mauricio Caicedo (Jefe SST Sacyr)
    {
        "id_contacto": "C011",
        "empresa": "Sacyr Construcción Colombia S.A.S.",
        "contacto": "Mauricio Caicedo",
        "cargo": "Jefe de Seguridad Industrial y SST",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Mauricio, con Sacyr operando 10+ proyectos en Colombia incluyendo 4 con el IDU de Bogotá, la gestión de permisos de trabajo en múltiples frentes simultáneos es crítica. SEÑAL digitaliza ATS, permisos y checklists HSE con firma en campo y trazabilidad para auditoría. ¿Le muestro cómo en 20 minutos?"
        ),
        "estado": "Pendiente",
    },
    # E005 - C012 Javier Fernández (PM Sacyr)
    {
        "id_contacto": "C012",
        "empresa": "Sacyr Construcción Colombia S.A.S.",
        "contacto": "Javier Fernández Caamaño",
        "cargo": "Project Manager",
        "canal": "LinkedIn",
        "asunto": None,
        "mensaje": (
            "Javier, coordinando proyectos de infraestructura de la magnitud de Sacyr Colombia, el control de permisos de trabajo y cumplimiento HSE en campo puede ser el mayor riesgo operativo. SEÑAL opera desde el celular sin infraestructura adicional y se integra a su flujo de obra. ¿Hablamos esta semana?"
        ),
        "estado": "Pendiente",
    },
    # E005 - Email genérico Sacyr
    {
        "id_contacto": "C011",
        "empresa": "Sacyr Construcción Colombia S.A.S.",
        "contacto": "Mauricio Caicedo",
        "cargo": "Jefe de Seguridad Industrial y SST",
        "canal": "Email",
        "asunto": "Gestión digital de permisos de trabajo — proyectos IDU Bogotá",
        "mensaje": (
            "Buenos días,\n\n"
            "Sacyr tiene 4 proyectos activos con el IDU de Bogotá y más de 10 en Colombia. A esa escala, el coordinador HSE no puede estar en todos los frentes verificando que cada permiso de trabajo esté firmado antes del inicio de actividad.\n\n"
            "SEÑAL digitaliza ese control: el trabajador firma en el celular, usted ve el estado en tiempo real desde cualquier lugar. Listo para auditorías del Ministerio y la ARL.\n\n"
            "¿Le agendamos 20 minutos esta semana?\n\n"
            "Cordialmente,\nEquipo SEÑAL — Kairos DLS Group S.A.S.\ncvillamil@sacyr.com"
        ),
        "estado": "Pendiente",
    },
]

# =============================================================================
# PASO 5: GENERAR EXCEL
# =============================================================================

def make_header_fill():
    return PatternFill("solid", fgColor="2E4057")

def make_header_font():
    return Font(bold=True, color="FFFFFF", size=10)

def make_header_alignment():
    return Alignment(horizontal="center", vertical="center", wrap_text=True)

def make_caliente_fill():
    return PatternFill("solid", fgColor="D9EAD3")

def make_tibio_fill():
    return PatternFill("solid", fgColor="FFF2CC")

def make_wrap():
    return Alignment(wrap_text=True, vertical="top")

def set_headers(ws, headers, col_widths=None):
    ws.append(headers)
    for col_idx, cell in enumerate(ws[1], 1):
        cell.font = make_header_font()
        cell.fill = make_header_fill()
        cell.alignment = make_header_alignment()
    ws.row_dimensions[1].height = 30
    if col_widths:
        for i, w in enumerate(col_widths, 1):
            ws.column_dimensions[get_column_letter(i)].width = w

def apply_row_color(ws, row_idx, clasificacion):
    if clasificacion == "caliente":
        fill = make_caliente_fill()
    elif clasificacion == "tibio":
        fill = make_tibio_fill()
    else:
        return
    for cell in ws[row_idx]:
        cell.fill = fill

def generate_excel():
    wb = Workbook()

    # -------------------------------------------------------------------------
    # PESTAÑA 1: Empresas
    # -------------------------------------------------------------------------
    ws1 = wb.active
    ws1.title = "Empresas"
    headers1 = [
        "ID", "Empresa", "NIT", "Sector", "Subsector", "Ciudad",
        "Tamaño estimado", "Web", "LinkedIn empresa",
        "Score (1-10)", "Clasificación", "Plan sugerido", "Fuente", "Contexto"
    ]
    col_widths1 = [6, 28, 16, 14, 26, 14, 32, 36, 36, 10, 12, 12, 28, 55]
    set_headers(ws1, headers1, col_widths1)

    emp_map = {e["id"]: e for e in EMPRESAS}

    for e in EMPRESAS:
        row = [
            e["id"], e["nombre"], e["nit"], e["sector"], e["subsector"], e["ciudad"],
            e["tamano"], e["web"], e["linkedin"] or "",
            e["score"], e["clasificacion"], e["plan"], e["fuente"], e["contexto"]
        ]
        ws1.append(row)
        row_idx = ws1.max_row
        apply_row_color(ws1, row_idx, e["clasificacion"])
        # URLs como hipervínculos
        ws1.cell(row=row_idx, column=8).hyperlink = e["web"]
        ws1.cell(row=row_idx, column=8).font = Font(color="0563C1", underline="single")
        if e["linkedin"]:
            ws1.cell(row=row_idx, column=9).hyperlink = e["linkedin"]
            ws1.cell(row=row_idx, column=9).font = Font(color="0563C1", underline="single")
        for col in range(1, len(headers1)+1):
            ws1.cell(row=row_idx, column=col).alignment = make_wrap()

    ws1.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 2: Contactos
    # -------------------------------------------------------------------------
    ws2 = wb.create_sheet("Contactos")
    headers2 = [
        "ID contacto", "ID empresa", "Empresa", "Nombre", "Cargo",
        "Prioridad cargo", "Email", "Confianza email", "LinkedIn", "Teléfono", "Fuente"
    ]
    col_widths2 = [10, 8, 28, 26, 32, 12, 36, 22, 55, 20, 28]
    set_headers(ws2, headers2, col_widths2)

    prioridad_label = {
        1: "1-HSE/SST", 2: "2-Operaciones", 3: "3-Gerente Gral",
        4: "4-CTO/Innovación", 5: "5-Gte Administrativo"
    }

    for c in CONTACTOS:
        e = emp_map.get(c["id_empresa"], {})
        clasificacion = e.get("clasificacion", "")
        row = [
            c["id"], c["id_empresa"], c["empresa"], c["nombre"], c["cargo"],
            prioridad_label.get(c["prioridad"], str(c["prioridad"])),
            c["email"] or "", c["confianza_email"] or "",
            c["linkedin"] or "", c["telefono"] or "", c["fuente"]
        ]
        ws2.append(row)
        row_idx = ws2.max_row
        apply_row_color(ws2, row_idx, clasificacion)
        if c["linkedin"]:
            ws2.cell(row=row_idx, column=9).hyperlink = c["linkedin"]
            ws2.cell(row=row_idx, column=9).font = Font(color="0563C1", underline="single")
        for col in range(1, len(headers2)+1):
            ws2.cell(row=row_idx, column=col).alignment = make_wrap()

    ws2.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 3: Mensajes
    # -------------------------------------------------------------------------
    ws3 = wb.create_sheet("Mensajes")
    headers3 = [
        "ID contacto", "Empresa", "Contacto", "Cargo",
        "Canal", "Asunto (email)", "Mensaje", "Estado envío"
    ]
    col_widths3 = [10, 28, 26, 32, 10, 55, 90, 14]
    set_headers(ws3, headers3, col_widths3)

    contacto_map = {c["id"]: c for c in CONTACTOS}

    for m in MENSAJES:
        c = contacto_map.get(m["id_contacto"], {})
        e = emp_map.get(c.get("id_empresa", ""), {})
        clasificacion = e.get("clasificacion", "")
        row = [
            m["id_contacto"], m["empresa"], m["contacto"], m["cargo"],
            m["canal"], m["asunto"] or "", m["mensaje"], m["estado"]
        ]
        ws3.append(row)
        row_idx = ws3.max_row
        apply_row_color(ws3, row_idx, clasificacion)
        for col in range(1, len(headers3)+1):
            ws3.cell(row=row_idx, column=col).alignment = make_wrap()
        # Ajustar altura de fila para mensajes largos
        ws3.row_dimensions[row_idx].height = 80

    ws3.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # PESTAÑA 4: Pipeline
    # -------------------------------------------------------------------------
    ws4 = wb.create_sheet("Pipeline")
    headers4 = [
        "ID empresa", "Empresa", "Score", "Clasificación", "Plan sugerido",
        "Contacto principal", "Cargo contacto", "Canal preferido",
        "Etapa", "Fecha primer contacto", "Próxima acción",
        "Responsable comercial", "Notas"
    ]
    col_widths4 = [10, 28, 8, 12, 12, 26, 32, 14, 16, 20, 32, 22, 45]
    set_headers(ws4, headers4, col_widths4)

    # Contacto principal por empresa (prioridad más alta encontrada)
    empresa_contacto_principal = {}
    for c in CONTACTOS:
        eid = c["id_empresa"]
        if eid not in empresa_contacto_principal:
            empresa_contacto_principal[eid] = c
        else:
            if c["prioridad"] < empresa_contacto_principal[eid]["prioridad"]:
                empresa_contacto_principal[eid] = c

    # Ordenar por score descendente
    empresas_ordenadas = sorted(EMPRESAS, key=lambda x: x["score"], reverse=True)

    for e in empresas_ordenadas:
        cp = empresa_contacto_principal.get(e["id"], {})
        # Canal preferido: LinkedIn si no hay email; Email si tiene
        canal_pref = "Email" if cp.get("email") else "LinkedIn"
        row = [
            e["id"], e["nombre"], e["score"], e["clasificacion"], e["plan"],
            cp.get("nombre", ""), cp.get("cargo", ""), canal_pref,
            "", "", "", "", ""  # Columnas vacías para el comercial
        ]
        ws4.append(row)
        row_idx = ws4.max_row
        apply_row_color(ws4, row_idx, e["clasificacion"])
        for col in range(1, len(headers4)+1):
            ws4.cell(row=row_idx, column=col).alignment = make_wrap()

    ws4.freeze_panes = "A2"

    # -------------------------------------------------------------------------
    # Guardar
    # -------------------------------------------------------------------------
    wb.save(OUTPUT_PATH)
    print(f"Archivo generado: {OUTPUT_PATH}")


if __name__ == "__main__":
    generate_excel()

    # Estadísticas finales
    print("\n=== REPORTE FINAL ===")
    print(f"Empresas encontradas: {len(EMPRESAS)}")
    print(f"Empresas en pipeline: {sum(1 for e in EMPRESAS if e['score'] >= 3)}")
    print(f"Contactos enriquecidos: {len(set(c['id'] for c in CONTACTOS))}")
    print(f"Mensajes generados: {len(MENSAJES)}")
    print("\nTop 3 leads por score:")
    top3 = sorted(EMPRESAS, key=lambda x: x['score'], reverse=True)[:3]
    for i, e in enumerate(top3, 1):
        # Buscar contacto principal
        cp = None
        best_prio = 99
        for c in CONTACTOS:
            if c['id_empresa'] == e['id'] and c['prioridad'] < best_prio:
                cp = c
                best_prio = c['prioridad']
        print(f"  {i}. {e['nombre']} | Score: {e['score']} | Clasificación: {e['clasificacion']} | Contacto: {cp['nombre'] if cp else 'N/A'} ({cp['cargo'] if cp else ''})")
